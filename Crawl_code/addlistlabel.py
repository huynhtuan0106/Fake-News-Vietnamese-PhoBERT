import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook

# Dữ liệu số dạng chuỗi
data = """[0 1 0 1 1 0 0 1 1 1 1 0 0 1 1 0 0 0 0 1 0 1 0 1 0 1 0 0 1 1 1 0 1 1 0 1 1
 0 0 0 1 1 1 0 1 1 0 0 1 1 0 0 0 0 0 0 1 0 0 1 0 1 0 1 1 1 0 0 0 1 0 0 1 0
 1 1 0 0 1 0 1 0 0 0 1 0 0 0 1 0 1 1 0 1 0 0 1 1 1 0 1 1 0 0 0 1 1 0 0 1 0
 0 1 0 1 1 0 1 1 0 1 0 1 1 1 0 0 0 1 1 0 0 0 1 1 0 0 1 1 0 0 0 0 1 1 1 1 1
 1 1 1 1 0 1 0 0 1 1 1 0 0 1 0 0 0 1 0 0 0 1 1 0 0 1 1 1 1 1 0 0 1 1 0 1 1
 1 1 1 1 1 1 1 1 0 1 1 1 1 0 0 1 0 1 0 0 0 1 1 0 0 1 0 0 1 1 1 1 0 0 1 1 0
 1 0 0 1 1 1 0 1 1 0 0 1 0 0 0 0 1 0 1 0 1 1 0 1 1 1 1 1 1 0 1 1 0 0 0 0 1
 1 0 0 0 1 1 1 1 0 1 1 0 0 0 1 1 0 1 1 1 1 0 1 1 0 0 1 0 1 0 0 1 1 1 1 0 0
 0 1 0 0 1 0 1 0 0 1 0 0 1 0 0 1 0 0 1 0 1 1 1 0 1 0 1 1 0 1 1 1 1 0 1 1 1
 1 0 1 0 0 1 1 1 0 0 1 1 1 0 1 0 1 1 1 0 1 1 1 0 1 1 1 0 0 0 1 1 1 1 0 1 0
 1 0 0 0 1 0 0 0 0 1 1 1 1 1 1 0 0 0 0 1 0 1 0 0 1 0 0 0 0 0 0 0 0 0 0 0 0
 0 1 1 0 1 0 0 0 1 0 0 0 1 0 1 1 1 0 0 0 1 0 0 1 1 0 1 1 1 1 1 0 1 1 1 0 1
 1 0 1 0 1 0 1 1 0 1 0 0 1 1 1 0 1 0 1 1 0 0 1 0 0 1 0 0 0 0 0 0 1 0 0 0 1
 1 0 0 0 0 0 0 1 0 1 0 1 0 0 0 1 0 0 0 0]"""

# Chuyển chuỗi thành danh sách các số nguyên
data_list = data.strip("[]").split()
data_list = [int(num) for num in data_list]

# Mở workbook từ file Excel hiện có
wb = load_workbook('Data/Draft/url.xlsx')
ws = wb.active  # Hoặc chọn worksheet cụ thể bằng tên nếu cần: wb['Tên worksheet']

# Ghi dữ liệu vào cột A, bắt đầu từ hàng trống đầu tiên
start_row = ws.max_row + 1  # Xác định hàng trống tiếp theo
for index, value in enumerate(data_list, start=start_row):
    ws[f'F{index}'] = value

# Lưu workbook sau khi thêm dữ liệu
wb.save('Data/Draft/url.xlsx')

print("Dữ liệu đã được thêm vào file 'data/draft/url.xlsx'.")
